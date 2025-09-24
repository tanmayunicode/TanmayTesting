import logging

import openpyxl
import pytest

@pytest.mark.usefixtures("startUP")
class BaseClass:
    print("I am in BaseClass")

    def generateLogs(self):
        logger = logging.getLogger(__name__)
        filehandler = logging.FileHandler("E:/Batches/PythonFramework/BA275PythonFramework/logs/MyLog.log")
        fileFormatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
        filehandler.setFormatter(fileFormatter)
        logger.addHandler(filehandler)
        logger.setLevel(logging.DEBUG)
        return logger

    @staticmethod
    def getTestData(testcasename):
        list = []
        tup = ()
        MyList = []

        excelfile = openpyxl.load_workbook(
            "E:/Batches/PythonFramework/BA275PythonFramework/testdata/TestData.xlsx")
        allSheets = excelfile.sheetnames

        for sheetname in allSheets:
            if sheetname == testcasename:
                sheet = excelfile[sheetname]

        rows = sheet.max_row
        cols = sheet.max_column

        for r in range(2, rows + 1):
            for c in range(1, cols + 1):
                list.append(str(sheet.cell(row=r, column=c).value))

            tup = tuple(list)
            list = []
            MyList.append(tup)

        return MyList

